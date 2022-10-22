import openpyxl
import pprint

# Open and read the Workbook
print('Opening workbook...')
wb = openpyxl.load_workbook('./data/notas.xlsx')
sheet = wb['1°E']
alumnoData = {}

GRADO = sheet['B3'].value.strip('°')
SECCION = sheet['B5'].value.upper()
CURSOSAPTITUDES = ['Desarrollo Personal, Ciudadanía y Cívica', ['Construye su identidad', 'Convive y participa democráticamente en la búsqueda del bien común'], 'Ciencias Sociales', [
    'Construye interpretaciones históricas', 'Gestiona responsablemente el espacio y el ambiente', 'Gestiona responsablemente los recursos económicos']]

# TODO: Fill in alumnoData with each student info.
print('Reading rows...')
row = 10
while (True):
    name = sheet['B' + str(row)].value

    if name == None:
        break

    aptitud = sheet['C9'].value
    nota = sheet['C' + str(row)].value
    nroOrden = row - 9

    # Adds a new student to the dict
    alumnoData.setdefault(name, {})

    # TODO: Adds 'notas', more than course.
    # Adds 'grado', 'seccion', 'nroOrden' and 'cursos'
    alumnoData[name].setdefault('grado', GRADO)
    alumnoData[name].setdefault('seccion', SECCION)
    alumnoData[name].setdefault('nroOrden', nroOrden)
    alumnoData[name].setdefault('cursos', {})

    for curso in CURSOS:
        alumnoData[name]['cursos'].setdefault(curso, {})

    # alumnoData[name]['cursos'][curso].setdefault(aptitud, nota)

    print(row)
    row += 1


# Expected output for each student:
# alumnoData = {'AGUIRRE PRADO...': {'grado': 1, 'seccion': 'e',
#                                    'cursos': {'desarrollo...': {'aptitud1': 'A', 'aptitud2': 'B'}}}}


# print('B47: ' + str(sheet['B47'].value))
# print('C8: ' + sheet['C8'].value)
# print('D8: ' + sheet['C8'].value)
# print('GRADO: ' + str(GRADO))
# print('SECCION: ' + str(SECCION))
# print('max_row: ' + str(sheet.max_row))
# print('max_column: ' + str(sheet.max_column))
# print()
print(pprint.pformat(alumnoData))
